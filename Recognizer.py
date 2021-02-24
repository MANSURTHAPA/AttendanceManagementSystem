
from imutils.video import VideoStream
from imutils.video import FPS
import face_recognition
import imutils
import pickle
import time
import datetime
import pandas as pd
from df2gspread import df2gspread as d2g

import cv2
from openpyxl import Workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials

print("[INFO] starting video stream...")

vs=VideoStream(src=0).start()

scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name('attendancesystem-304808-db9eced2794f.json', scope)
client = gspread.authorize(credentials)
sheet2 = client.open("AttendanceSystem").worksheets()
book = Workbook()
sheet = book.active
spreadsheetkey='1bRa0NqPdfX0IQVarLE_DC4Lb0eevawp1Vs0pKUrzaj0'
wks='Sheet1'
# Initialize 'currentname' to trigger only when a new person is identified.
currentname = "unknown"
# Determine faces from encodings.pickle file model created from train_model.py
encodingsP = "encodings.pickle"
# use this xml file
cascade = "haarcascade_frontalface_default.xml"
currentDate = time.strftime("%d_%m_%y")

# load the known faces and embeddings along with OpenCV's Haar
# cascade for face detection
print("[INFO] loading encodings + face detector...")
data = pickle.loads(open(encodingsP, "rb").read())
detector = cv2.CascadeClassifier(cascade)

# initialize the video stream and allow the camera sensor to warm up

# vs = VideoStream(usePiCamera=True).start()
now = datetime.datetime.now()
today = now.day
month = now.month
time.sleep(2.0)

f = open("datatext.csv", "r+")
studnet = f.readlines()
regNumberList = []
stdname=[]
sname=""



for line in studnet:
    entry = line.split(' ')
    regNumberList.append(entry[0])
    stdname.append(entry[1])

print(regNumberList)
print(regNumberList)


# start the FPS counter
fps = FPS().start()
for i in range(len(regNumberList)):
    sheet.cell(row=int(i + 1), column=int(1)).value = currentDate
    sheet.cell(row=int(i + 1), column=int(2)).value = stdname[i]
    sheet.cell(row=int(i + 1), column=int(3)).value = "Absent"
    book.save(str(month) + '.xlsx')
# loop over frames from the video file stream



# for ii in sheet2:
#     print(ii.title)
#     curr_time=time.localtime()
#     time_str=time.strftime("%m/%d/%Y  %H:%M:%S",curr_time)
#     ii.append_row([time_str,sname,"Absnet"])
while True:
    # grab the frame from the threaded video stream and resize it
    # to 500px (to speedup processing)
    frame = vs.read()
    frame = imutils.resize(frame, width=500)

    # convert the input frame from (1) BGR to grayscale (for face
    # detection) and (2) from BGR to RGB (for face recognition)
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # detect faces in the grayscale frame
    rects = detector.detectMultiScale(gray, scaleFactor=1.1,
                                      minNeighbors=5, minSize=(30, 30),
                                      flags=cv2.CASCADE_SCALE_IMAGE)

    # OpenCV returns bounding box coordinates in (x, y, w, h) order
    # but we need them in (top, right, bottom, left) order, so we
    # need to do a bit of reordering
    boxes = [(y, x + w, y + h, x) for (x, y, w, h) in rects]

    # compute the facial embeddings for each face bounding box
    encodings = face_recognition.face_encodings(rgb, boxes)
    names = []
    counts = {}

    # loop over the facial embeddings
    for encoding in encodings:
        # attempt to match each face in the input image to our known
        # encodings
        matches = face_recognition.compare_faces(data["encodings"],
                                                 encoding)
        name = "Unknown"  # if face is not recognized, then print Unknown

        # check to see if we have found a match
        if True in matches:

            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            #
            #
            #
            # # loop over the matched indexes and maintain a count for
            # # each recognized face face
            for i in matchedIdxs:
                name = data["names"][i]
                id = data["id"][i]
                counts[name, id] = counts.get((name, id), 0) + 1

            myname = max(counts, key=counts.get)
            print(myname)



            if id in regNumberList:
                sheet.cell(row=int(id), column=int(1)).value = currentDate
                sheet.cell(row=int(id), column=int(2)).value = name
                sheet.cell(row=int(id), column=int(3)).value = "Present"
                book.save(str(month) + '.xlsx')
            df = pd.read_excel(r'./2.xlsx',sheet_name=0,header=None,usecols=[0,1,2],names=['Time','Name','Attendance'],dtype={2:str})
            print(df)

            d2g.upload(df, spreadsheetkey, wks, credentials=credentials, row_names=True)


            # If someone in your dataset is identified, print their name on the screen
            if currentname != name:
                currentname = name
                print(currentname)
                break

            # determine the recognized face with the largest number
            # of votes (note: in the event of an unlikely tie Python
            # will select first entry in the dictionary)

        # update the list of names
        names.append(name)

    # loop over the recognized faces
    for ((top, right, bottom, left), name) in zip(boxes, names):
        # draw the predicted face name on the image - color is in BGR
        cv2.rectangle(frame, (left, top), (right, bottom),
                      (0, 255, 225), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,
                    .8, (0, 255, 255), 2)

    # display the image to our screen
    cv2.imshow("Facial Recognition is Running", frame)

    key = cv2.waitKey(1) & 0xFF

    # quit when 'q' key is pressed
    if key == 27 or key == ord("q"):
        break

    # update the FPS counter
    fps.update()

# stop the timer and display FPS information
fps.stop()
print("[INFO] elasped time: {:.2f}".format(fps.elapsed()))
print("[INFO] approx. FPS: {:.2f}".format(fps.fps()))

# do a bit of cleanup
cv2.destroyAllWindows()
vs.stop()
